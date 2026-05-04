"""
Ingest script for QuanFeng Smart Brain RAG system.
Processes PDF and Excel files, chunks them, and stores embeddings in ChromaDB.
"""
import os
import uuid
import chromadb
from chromadb.config import Settings
import openpyxl
import pdfplumber

# Configuration
DATA_DIR = r"C:\Users\III-AIPC-02\.nanobot_2\workspace\quanfeng_brain\data"
VECTOR_STORE_DIR = r"C:\Users\III-AIPC-02\.nanobot_2\workspace\quanfeng_brain\vector_store"
COLLECTION_NAME = "quanfeng_docs"

# Ollama settings
EMBEDDING_MODEL = "nomic-embed-text"  # Good for Chinese/English
CHUNK_SIZE = 500
CHUNK_OVERLAP = 50


def extract_pdf_text(pdf_path: str) -> str:
    """Extract text from PDF using pdfplumber."""
    full_text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text()
            if text:
                full_text += f"\n--- Page {i+1} ---\n{text}\n"
    return full_text


def extract_excel_text(excel_path: str) -> str:
    """Extract text from Excel FAQ file."""
    full_text = ""
    wb = openpyxl.load_workbook(excel_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        full_text += f"\n--- Sheet: {sheet_name} ---\n"
        for row in sheet.iter_rows(values_only=True):
            # Filter out None values and join
            row_text = " | ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                full_text += row_text + "\n"
    return full_text


def chunk_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list:
    """Split text into overlapping chunks."""
    chunks = []
    # Split by paragraphs first
    paragraphs = text.split("\n")
    current_chunk = ""
    
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        
        if len(current_chunk) + len(para) > chunk_size:
            if current_chunk:
                chunks.append(current_chunk)
            current_chunk = para
        else:
            if current_chunk:
                current_chunk += "\n" + para
            else:
                current_chunk = para
    
    if current_chunk:
        chunks.append(current_chunk)
    
    return chunks


def create_embeddings(chunks: list, source_file: str) -> list:
    """Create documents with metadata for ChromaDB."""
    documents = []
    for i, chunk in enumerate(chunks):
        doc = {
            "id": f"{source_file}_chunk_{i}",
            "text": chunk,
            "metadata": {
                "source": source_file,
                "chunk_index": i
            }
        }
        documents.append(doc)
    return documents


def ingest_to_chroma(documents: list):
    """Store documents in ChromaDB using Ollama embeddings."""
    # Initialize ChromaDB client
    client = chromadb.PersistentClient(path=VECTOR_STORE_DIR)
    
    # Get or create collection
    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"}
    )
    
    # Prepare data for ChromaDB
    ids = [doc["id"] for doc in documents]
    texts = [doc["text"] for doc in documents]
    metadatas = [doc["metadata"] for doc in documents]
    
    # Add to collection (ChromaDB will use Ollama for embeddings if configured)
    # Note: We need to set up Ollama embedding function
    import ollama
    
    # Generate embeddings using Ollama
    print(f"Generating embeddings for {len(texts)} chunks using Ollama...")
    embeddings = []
    for text in texts:
        response = ollama.embed(model=EMBEDDING_MODEL, input=text)
        embeddings.append(response["embeddings"][0])
    
    # Add to collection
    collection.add(
        ids=ids,
        documents=texts,
        metadatas=metadatas,
        embeddings=embeddings
    )
    
    print(f"Successfully ingested {len(documents)} chunks into ChromaDB.")


def main():
    # Create directories
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(VECTOR_STORE_DIR, exist_ok=True)
    
    # Source files
    pdf_path = r"D:\nanodemo\全鋒\國泰金控 20260424 (162大車補貼參數調整).pdf"
    excel_path = r"D:\nanodemo\全鋒\道路救援通則FAQ.xlsx"
    
    all_documents = []
    
    # Process PDF
    if os.path.exists(pdf_path):
        print("Processing PDF...")
        pdf_text = extract_pdf_text(pdf_path)
        pdf_chunks = chunk_text(pdf_text)
        pdf_docs = create_embeddings(pdf_chunks, "國泰金控_大車補貼.pdf")
        all_documents.extend(pdf_docs)
        print(f"  -> {len(pdf_chunks)} chunks from PDF")
    
    # Process Excel
    if os.path.exists(excel_path):
        print("Processing Excel...")
        excel_text = extract_excel_text(excel_path)
        excel_chunks = chunk_text(excel_text)
        excel_docs = create_embeddings(excel_chunks, "道路救援通則FAQ.xlsx")
        all_documents.extend(excel_docs)
        print(f"  -> {len(excel_chunks)} chunks from Excel")
    
    # Ingest to ChromaDB
    if all_documents:
        print("\nIngesting to ChromaDB...")
        ingest_to_chroma(all_documents)
    else:
        print("No documents to ingest.")


if __name__ == "__main__":
    main()
